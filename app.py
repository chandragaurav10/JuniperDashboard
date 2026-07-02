import os

import streamlit as st
import pandas as pd
import plotly.express as px
import matplotlib

from io import BytesIO
from openpyxl import Workbook
from openpyxl.styles import (
    PatternFill,
    Font,
    Border,
    Side
)
from openpyxl.styles import Border, Side
from openpyxl.utils import get_column_letter


from selenium import webdriver
from selenium.webdriver.chrome.options import Options

import time

from reportlab.platypus import (
    SimpleDocTemplate,
    Paragraph,
    Spacer
)

from reportlab.lib.styles import getSampleStyleSheet

import gspread


# ==========================
# Just for local testing, comment this section when deploying to Streamlit Cloud
# ==========================

#try:
#    gc = gspread.service_account_from_dict(
#        st.secrets["gcp_service_account"]
#    )

#except:
#    gc = gspread.service_account(
#        filename="juniper-daily-sale-0bf842c78f6c.json"
#    )


# ==========================
# Commenting for time being to avoid gspread error in local environment
# ==========================

gc = gspread.service_account_from_dict(
   st.secrets["gcp_service_account"]
)

sheet = gc.open("Juniper Project File")

worksheet = sheet.worksheet("Sales Data Non-Air 2026")

data = worksheet.get_all_records()

df = pd.DataFrame(data)




df.columns = df.columns.str.strip()

df["Agency Type"] = (
    df["Agency Type"]
    .astype(str)
    .str.strip()
)

df["Agency Type"] = (
    df["Agency Type"]
    .replace("", "Unknown")
    .fillna("Unknown")
)

df["Branch Name"] = (
    df["Branch Name"]
    .astype(str)
    .str.strip()
)

# ==========================
# BOOKING STATUS
# ==========================

df["Booking Status"] = (
    df["Booking Status"]
    .astype(str)
    .str.strip()
    .replace("", "Unknown")
    .fillna("Unknown")
)

# ==========================
# PAGE CONFIG
# ==========================

st.set_page_config(
    page_title="Juniper Sales Dashboard",
    page_icon="📈",
    layout="wide"
)
st.markdown("""
<style>

[data-testid="stSidebar"]{
    background:#f8fafc;
    border-right:2px solid #e5e7eb;
}

.stMultiSelect div[data-baseweb="select"]{
    border-radius:12px;
}

.stSelectbox div[data-baseweb="select"]{
    border-radius:12px;
}

.stDateInput{
    border-radius:12px;
}

</style>
""", unsafe_allow_html=True)

pdf_mode = st.query_params.get("pdf", "0") == "1"
if pdf_mode:
    st.markdown("""
    <style>

    section.main > div{
        max-width:100% !important;
        padding-left:1rem !important;
        padding-right:1rem !important;
    }

    [data-testid="stSidebar"]{
        display:none !important;
    }

    .pdf-page{
        page-break-after: always;
        break-after: page;
        min-height:100vh;
    }

    </style>
    """, unsafe_allow_html=True)

st.markdown("""
<style>

[data-testid="stSidebar"] {
    background-color: #f8fafc;
}

[data-testid="stSidebar"] h1,
[data-testid="stSidebar"] h2,
[data-testid="stSidebar"] h3 {
    color: #0f172a;
}

[data-testid="stSidebar"] {
    border-right: 2px solid #e2e8f0;
}

</style>
""", unsafe_allow_html=True)


if pdf_mode:
    st.markdown("""
<style>

[data-testid="stSidebar"]{
    display:none !important;
}

.stAlert{
    display:none !important;
}

.stButton{
    display:none !important;
}

</style>
""", unsafe_allow_html=True)

# ==========================
# CLEAN DATA
# ==========================

df["Sale'26"] = pd.to_numeric(df["Sale'26"], errors="coerce")
df.rename(columns={"Sale'26": "Sale (€)"}, inplace=True)
df["Month Number"] = pd.to_numeric(df["Month Number"], errors="coerce")

# ==========================
# KPIs
# ==========================

df["Date Convert"] = pd.to_datetime(
    df["Date Convert"],
    dayfirst=True,
    errors="coerce"
)

# ==========================
# SIDEBAR FILTERS
# ==========================

st.sidebar.markdown("""
<div style="
background: linear-gradient(135deg,#1e293b,#334155);
padding:15px;
border-radius:12px;
margin-bottom:15px;
text-align:center;
">
<h2 style="color:white;margin:0;">
🎛 Dashboard Filters
</h2>
</div>
""", unsafe_allow_html=True)

selected_branch = st.sidebar.multiselect(
    "Branch",
    sorted(df["Branch Name"].dropna().unique())
)

month = st.sidebar.multiselect(
    "Month",
    options=sorted(df["Month Number"].dropna().unique()),
    default=sorted(df["Month Number"].dropna().unique())
)

st.sidebar.markdown("### 📅 Date Range")

# =========================
# DATE RANGE FILTER
# =========================

df["Date Convert"] = pd.to_datetime(
    df["Date Convert"],
    dayfirst=True,
    errors="coerce"
)

valid_dates = df["Date Convert"].dropna()

if valid_dates.empty:
    st.error("No valid dates found in 'Date Convert' column.")
    st.stop()

min_date = valid_dates.min().date()
max_date = valid_dates.max().date()

from_date = st.sidebar.date_input(
    "From Date",
    value=min_date,
    min_value=min_date,
    max_value=max_date
)

to_date = st.sidebar.date_input(
    "To Date",
    value=max_date,
    min_value=min_date,
    max_value=max_date
)

# ==========================
# DESTINATION FILTER
# ==========================

st.sidebar.markdown("### 🌍 Destination")

destination_list = sorted(
    df["Product's Country"]
    .dropna()
    .unique()
)

selected_destination = st.sidebar.selectbox(
    "Destination",
    ["All Destinations"] + destination_list
)

# ==========================
# BOOKING STATUS FILTER
# ==========================

st.sidebar.markdown("### 📌 Booking Status")

booking_status = st.sidebar.multiselect(
    "Booking Status",
    options=sorted(df["Booking Status"].unique()),
    default=sorted(df["Booking Status"].unique())
)

agency_filter = st.session_state.get("agency_filter", "All")

filtered_df = df[
    df["Month Number"].isin(month)
]

if booking_status:
    filtered_df = filtered_df[
        filtered_df["Booking Status"].isin(booking_status)
    ]

if agency_filter != "All":
    filtered_df = filtered_df[
        filtered_df["Agency Type"] == agency_filter
    ]

if selected_branch:
    filtered_df = filtered_df[
        filtered_df["Branch Name"].isin(selected_branch)
    ]

# Remove invalid dates before applying date filter
filtered_df = filtered_df.dropna(subset=["Date Convert"])

filtered_df = filtered_df[
    (filtered_df["Date Convert"].dt.date >= from_date)
    &
    (filtered_df["Date Convert"].dt.date <= to_date)
]

import datetime

# KPI values from COMPLETE sheet

# Convert safely
df["Date Convert"] = pd.to_datetime(
    df["Date Convert"],
    dayfirst=True,
    errors="coerce"
)
# Remove invalid dates
valid_dates = df["Date Convert"].dropna()
if valid_dates.empty:
    latest_date = None
else:
    latest_date = valid_dates.dt.date.max()

# Last Day Sale = yesterday
yesterday = datetime.date.today() - datetime.timedelta(days=1)
daily_sale = filtered_df.loc[
    filtered_df["Date Convert"].dt.date == yesterday,
    "Sale (€)"
].sum()

# Today's Sale
today = datetime.date.today()
today_sale = filtered_df.loc[
    filtered_df["Date Convert"].dt.date == today,
    "Sale (€)"
].sum()


# Timestamp of last data push today
from datetime import datetime
import pytz

ist = pytz.timezone("Asia/Kolkata")

last_updated = datetime.now(ist).strftime(
    "%d %b %Y, %I:%M %p"
)
#last_updated = now.strftime("%d %b %Y, %I:%M %p")


# MTD = current month only
current_month = today.month
current_year  = today.year
mtd_sale = filtered_df.loc[
    (filtered_df["Date Convert"].dt.month == current_month) &
    (filtered_df["Date Convert"].dt.year  == current_year),
    "Sale (€)"
].sum()

# Total Sale = full filtered dataset
ytd_sale = filtered_df["Sale (€)"].sum()

latest_month = df["Month Number"].max()

bookings = len(filtered_df)

# ====================================
# DASHBOARD PDF FUNCTION
# ====================================


def create_dashboard_pdf():

    chrome_options = Options()

    chrome_options.add_argument("--headless=new")
    chrome_options.add_argument("--window-size=1920,12000")

    driver = webdriver.Chrome(options=chrome_options)

    driver.get("http://localhost:8501/?pdf=1")

    time.sleep(10)

    if "can't reach this page" in driver.page_source.lower():
        raise Exception(
        "Streamlit is not running on localhost:8501"
    )

    driver.maximize_window()

    time.sleep(10) # Wait for the page to fully render
    height = driver.execute_script(""" return document.documentElement.scrollHeight""")
    driver.set_window_size(1920, height)
    time.sleep(2) 
    pdf_data = driver.execute_cdp_cmd(
    "Page.printToPDF",
    {
        "printBackground": True,
        "landscape": True,
        "paperWidth": 11.69,
        "paperHeight": 11.69,
        "marginTop": 0,
        "marginBottom": 0,
        "marginLeft": 0,
        "marginRight": 0,
        "scale": 1
    }
)
    print("Page Height =", height)
    


    driver.quit()

    import base64

    pdf_bytes = base64.b64decode(
        pdf_data["data"]
    )

    return pdf_bytes


if not pdf_mode:

    st.sidebar.markdown("---")

    if st.sidebar.button(
        "📄 Generate Dashboard PDF"
    ):

        st.sidebar.success(
            "Generating PDF... Please wait 30 sec"
        )

        pdf_file = create_dashboard_pdf()

        st.sidebar.download_button(
            "⬇ Download Dashboard PDF",
            data=pdf_file,
            file_name="Juniper_Dashboard.pdf",
            mime="application/pdf"
        )

# ==========================
# HEADER
# ==========================

if pdf_mode:
    st.markdown('<div class="pdf-page">', unsafe_allow_html=True)
    
st.title("📈 Juniper Sales Dashboard")

st.markdown("""
<style>
.main .block-container {
    max-width: 95%;
}
</style>
""", unsafe_allow_html=True)

st.markdown("""
<style>
.kpi-card {
    background: white;
    padding: 20px;
    border-radius: 12px;
    border: 1px solid #e5e7eb;
    text-align: center;
    box-shadow: 0 2px 8px rgba(0,0,0,0.08);
}
.kpi-title {
    font-size: 16px;
    font-weight: 700;
    color: #555;
}
.kpi-value {
    font-size: 34px;
    font-weight: 800;
    color: #0f172a;
}
</style>
""", unsafe_allow_html=True)

col1, col2, col3, col4, col5 = st.columns(5)

with col1:
    st.markdown(f"""
    <div class="kpi-card">
        <div class="kpi-title">🕐 Today's Sale</div>
        <div class="kpi-value">€{today_sale:,.0f}</div>
        <div style="font-size:11px;color:#94a3b8;margin-top:6px;">🔄 {last_updated}</div>
    </div>
    """, unsafe_allow_html=True)

with col2:
    st.markdown(f"""
    <div class="kpi-card">
        <div class="kpi-title">Last Day Sale</div>
        <div class="kpi-value">€{daily_sale:,.0f}</div>
        <div style="font-size:11px;color:#94a3b8;margin-top:6px;">Yesterday</div>
    </div>
    """, unsafe_allow_html=True)

with col3:
    st.markdown(f"""
    <div class="kpi-card">
        <div class="kpi-title">MTD Sale</div>
        <div class="kpi-value">€{mtd_sale:,.0f}</div>
        <div style="font-size:11px;color:#94a3b8;margin-top:6px;">This month</div>
    </div>
    """, unsafe_allow_html=True)

with col4:
    st.markdown(f"""
    <div class="kpi-card">
        <div class="kpi-title">Total Sale</div>
        <div class="kpi-value">€{ytd_sale:,.0f}</div>
        <div style="font-size:11px;color:#94a3b8;margin-top:6px;">All time</div>
    </div>
    """, unsafe_allow_html=True)

with col5:
    st.markdown(f"""
    <div class="kpi-card">
        <div class="kpi-title">Bookings</div>
        <div class="kpi-value">{bookings:,}</div>
        <div style="font-size:11px;color:#94a3b8;margin-top:6px;">Total orders</div>
    </div>
    """, unsafe_allow_html=True)

st.divider()


# ==========================
# MONTHLY SALES TREND
# ==========================

st.markdown("## 📊 Monthly Sales Trend")

monthly_sale = (
    filtered_df.groupby("Month Number")["Sale (€)"]
    .sum()
    .reset_index()
    .sort_values("Month Number")
)

month_map = {
    1: "Jan",
    2: "Feb",
    3: "Mar",
    4: "Apr",
    5: "May",
    6: "Jun",
    7: "Jul",
    8: "Aug",
    9: "Sep",
    10: "Oct",
    11: "Nov",
    12: "Dec"
}

monthly_sale["Month"] = monthly_sale["Month Number"].map(month_map)

fig = px.bar(
    monthly_sale,
    x="Month",
    y="Sale (€)",
    text="Sale (€)",
    title="Monthly Sales Trend"
)

fig.update_traces(
    texttemplate='€%{text:,.0f}',
    textposition='outside',
    textfont=dict(
        size=13,
        color='black',
        family='Arial Black'
    ),
    cliponaxis=False
)

fig.update_layout(
    template="plotly_white",
    height=500,
    margin=dict(l=50, r=80, t=50, b=80),
    bargap=0.35,
    font=dict(
        size=13,
        color="black"
    )
)

fig.update_layout(uniformtext_minsize=10, uniformtext_mode='hide')

fig.update_xaxes(
    automargin=True,
    fixedrange=True,
    tickfont=dict(
        size=14,
        color="black"
    )
)


fig.update_yaxes(
    fixedrange=True,
    range=[0, monthly_sale["Sale (€)"].max() * 1.35]
)


st.plotly_chart(
    fig,
    use_container_width=True,
    
)

if pdf_mode:
    st.markdown("</div>", unsafe_allow_html=True)

agency_sale = (
    filtered_df.groupby("Agency Type")["Sale (€)"]
    .sum()
    .reset_index()
)


# ===========================
# BRANCH SALE               
# ==========================


#st.markdown("## 🏢 Branch Sale")

branch_sale = (
    filtered_df.groupby("Branch Name")["Sale (€)"]
    .sum()
    .reset_index()
    .sort_values("Sale (€)", ascending=False)
    .head(10)
)

fig_branch = px.bar(
    branch_sale,
    x="Branch Name",
    y="Sale (€)",
    text="Sale (€)",
    title="Top 10 Branches"
)

fig_branch.update_traces(
    texttemplate='€%{text:,.0f}',
    textposition='outside',
    textfont=dict(
        size=13,
        color='black',
        family='Arial Black'
    ),
    cliponaxis=False
)

fig_branch.update_layout(
    template="plotly_white",
    height=650,
    showlegend=False,
    margin=dict(l=50, r=80, t=50, b=80),
    bargap=0.25,
    font=dict(
        size=13,
        color="black"
    )
)

fig_branch.update_xaxes(
    tickfont=dict(
        size=13,
        color="black"
    )
)

fig_branch.update_yaxes(
    tickfont=dict(
        size=13,
        color="black"
    )
)

# ==========================
# BRANCH & AGENCY SALE
# ==========================

if pdf_mode:
    st.markdown('<div class="pdf-page">', unsafe_allow_html=True)

st.markdown("## 🧭 Branch & Agency Sale")



if not pdf_mode:

    agency_filter = st.radio(
        "Agency Type",
        ["All", "B2B", "HQ"],
        horizontal=False,
        key="agency_filter"
    )
else:
    agency_filter = "All"

fig_branch.update_layout(
    template="plotly_white",
    height=400,
    margin=dict(l=20, r=20, t=40, b=40)
)

st.plotly_chart(
    fig_branch,
    use_container_width=True
)


# ==========================
# DONUT CHART
# ==========================

total_agency_sale = ytd_sale

fig_agency = px.pie(
    agency_sale,
    names="Agency Type",
    values="Sale (€)",
    hole=0.65,
    color="Agency Type",
    color_discrete_map={
        "B2B": "#636EFA",
        "HQ": "#EF553B"
    }
)

fig_agency.update_traces(
    textinfo="percent",
    textfont_size=18,
    textfont_color="white",
    hovertemplate=
    "%{label}<br>" +
    "Sale: €%{value:,.0f}<br>" +
    "Share: %{percent}"
)

fig_agency.update_layout(
    template="plotly_white",
    height=400,
    showlegend=True,
    margin=dict(l=20, r=20, t=20, b=20),
    annotations=[
        dict(
            text=f"€{total_agency_sale:,.0f}",
            x=0.5,
            y=0.5,
            font_size=30,
            showarrow=False,
            font_color="#000000"
        )
    ]
)

b2b_sale = agency_sale.loc[
    agency_sale["Agency Type"] == "B2B",
    "Sale (€)"
].sum()

hq_sale = agency_sale.loc[
    agency_sale["Agency Type"] == "HQ",
    "Sale (€)"
].sum()

col1, col2 = st.columns([3,1])

with col1:
    st.plotly_chart(
        fig_agency,
        use_container_width=True
    )

with col2:
    st.metric(
        "🔵 B2B Sale",
        f"€{b2b_sale:,.0f}"
    )

    st.metric(
        "🔴 HQ Sale",
        f"€{hq_sale:,.0f}"
    )

if pdf_mode:
    st.markdown("</div>", unsafe_allow_html=True)




# ==========================
# TOP 10 DESTINATIONS
# ==========================

st.markdown("## 🌍 Top 10 Destinations")

destination_sale = (
    filtered_df.groupby("Product's Country")["Sale (€)"]
    .sum()
    .reset_index()
    .sort_values("Sale (€)", ascending=False)
    .head(10)
)

fig_destination = px.bar(
    destination_sale,
    x="Product's Country",
    y="Sale (€)",
    text="Sale (€)",
    title="Top 10 Destinations"
)

fig_destination.update_traces(
    texttemplate='€%{text:,.0f}',
    textposition='outside',
    textfont=dict(
        size=13,
        color='black',
        family='Arial Black'
    ),
    cliponaxis=False
)

fig_destination.update_layout(
    template="plotly_white",
    height=400,
    showlegend=False,
    margin=dict(l=50, r=80, t=50, b=80),
    bargap=0.25,
    font=dict(
        size=13,
        color="black"
    )
)

fig_destination.update_xaxes(
    tickangle=-20,
    tickfont=dict(
        size=13,
        color="black"
    )
)

fig_destination.update_yaxes(
    title="Sale (€)",
    range=[
        0,
        destination_sale["Sale (€)"].max() * 1.20
    ]
)

st.plotly_chart(
    fig_destination,
    use_container_width=True
)




#####

destination_df = filtered_df.copy()

if selected_destination != "All Destinations":
    destination_df = destination_df[
        destination_df["Product's Country"]
        == selected_destination
    ]


destination_df["Room Nights"] = (
    destination_df["No. of nights"]
    * destination_df["No. of Rooms"]
)



# ==========================
# TOP HOTELS BY TOTAL ROOM NIGHTS
# ==========================

st.markdown("## 🏨 Top 10 Hotels by Total Room Nights")

# Calculate Room Nights
filtered_df = filtered_df.copy()
filtered_df["Room Nights"] = (
    pd.to_numeric(
        filtered_df["No. of nights"],
        errors="coerce"
    ).fillna(0)
    *
    pd.to_numeric(
        filtered_df["No. of Rooms"],
        errors="coerce"
    ).fillna(0)
)

hotel_sale = (
    destination_df[
        destination_df["Description"].notna()
    ]
    .groupby("Description")["Room Nights"]
    .sum()
    .reset_index()
    .sort_values("Room Nights", ascending=False)
    .head(10)
)
hotel_sale["Hotel Name"] = hotel_sale["Description"]

fig_hotel = px.bar(
    hotel_sale,
    x="Hotel Name",
    y="Room Nights",
    text="Room Nights",
    title="Top 10 Hotels by Total Room Nights"
)

fig_hotel.update_traces(
    texttemplate='%{text:,.0f}',
    textposition='outside',
    textfont=dict(
        size=13,
        color='black',
        family='Arial Black'
    ),
    cliponaxis=False
)

fig_hotel.update_layout(
    template="plotly_white",
    height=400,
    xaxis_title="Hotel Name",
    showlegend=False,
    margin=dict(
        l=50,
        r=50,
        t=40,
        b=60
    ),
    bargap=0.25,
    font=dict(
        size=13,
        color="black"
    )
)

fig_hotel.update_xaxes(
    tickangle=-35,
    automargin=True,
    tickfont=dict(
        size=10,
        color="black"
    )
)

fig_hotel.update_yaxes(
    title="Total Room Nights",
    range=[
        0,
        hotel_sale["Room Nights"].max() * 1.20
    ]
)

st.plotly_chart(
    fig_hotel,
    use_container_width=True
)


# ==========================
# TOP 10 SUPPLIERS
# ==========================

#if pdf_mode:
#    st.markdown(
#        """
#        <div style="page-break-before: always;"></div>
#        """,
#        unsafe_allow_html=True
#    )

st.markdown("## 🏆 Top 10 Suppliers")



supplier_sale = (
    destination_df.groupby("Supplier Name")["Sale (€)"]
    .sum()
    .reset_index()
    .sort_values("Sale (€)", ascending=False)
    .head(10)
)

fig_supplier = px.bar(
    supplier_sale,
    x="Supplier Name",
    y="Sale (€)",
    text="Sale (€)",
    title="Top 10 Suppliers"
)

fig_supplier.update_traces(
    texttemplate='€%{text:,.0f}',
    textposition='outside',
    textfont=dict(
        size=13,
        color='black',
        family='Arial Black'
    ),
    cliponaxis=False
)

fig_supplier.update_layout(
    template="plotly_white",
    height=350,
    showlegend=False,
    margin=dict(l=40, r=40, t=40, b=80),
    bargap=0.25,
    font=dict(
        size=13,
        color="black"
    )
)

fig_supplier.update_xaxes(
    tickangle=-20,
    tickfont=dict(
        size=14,
        color="black"
    )
)

fig_supplier.update_yaxes(
    title="Sale (€)"
)

st.plotly_chart(
    fig_supplier,
    use_container_width=True
)







# TOP 5 NATIONALITIES
# ==========================

st.markdown("## 🌎 Top 5 Nationalities")

nationality_sale = (
    destination_df.groupby("Nationality")
    .size()
    .reset_index(name="Bookings")
    .sort_values("Bookings", ascending=False)
    .head(5)
)

fig_nat = px.bar(
    nationality_sale,
    x="Nationality",
    y="Bookings",
    text="Bookings",
    title=f"Top 5 Nationalities - {selected_destination}"
)

fig_nat.update_traces(
    texttemplate='%{text:,.0f}',
    textposition='outside',
    textfont=dict(
        size=13,
        color='black',
        family='Arial Black'
    ),
    cliponaxis=False
)

fig_nat.update_layout(
    template="plotly_white",
    height=350,
    showlegend=False,
    margin=dict(
        l=50,
        r=50,
        t=50,
        b=60
    ),
    bargap=0.25,
    font=dict(
        size=13,
        color="black"
    )
)

fig_nat.update_xaxes(
    tickangle=0,
    tickfont=dict(
        size=14,
        color="black"
    )
)

fig_nat.update_yaxes(
    tickfont=dict(
        size=13,
        color="black"
    )
)

st.plotly_chart(
    fig_nat,
    use_container_width=True
)


# ==========================
# BRANCH Daily Sale MATRIX
# ==========================

st.markdown("---")
st.markdown("## 📊 Daily Sales Report")

months = sorted(
    df["Month Number"]
    .dropna()
    .unique()
)
current_month = datetime.today().month

default_index = (
    months.index(current_month)
    if current_month in months
    else 0
)

col1, col2, col3 = st.columns([5,1,6])

with col1:
    report_type = st.radio(
        "Report Type",
        ["ALL", "HQ", "B2B"],
        horizontal=True,
        key="daily_sales_report"
    )

with col2:
    selected_month = st.selectbox(
        "Month",
        months,
        index=default_index
    )


# Filter Data Frame based on report type
monthly_df = filtered_df.copy()

monthly_df = monthly_df[
    monthly_df["Month Number"] == selected_month
]

if report_type == "HQ":
    monthly_df = monthly_df[
        monthly_df["Agency Type"] == "HQ"
    ]

elif report_type == "B2B":
    monthly_df = monthly_df[
        monthly_df["Agency Type"] == "B2B"
    ]



# MONTH FILTER
monthly_df = monthly_df[
    monthly_df["Month Number"] == selected_month
]


# Branch Short Name

branch_short = {
    "STTS ALA": "ALA",
    "STTS SVO": "SVO",
    "STTS AMM": "AMM",
    "STTS MAD": "MAD",
    "STTS CMN": "CMN",
    "STTS CDG": "CDG",
    "STTS DKR": "DKR",
    "STTS DOH": "DOH",
    "STTS BAH": "BAH",
    "STTS RUH": "RUH",
    "STTS CAN": "CAN",
    "STTS RAI": "RAI",
    "STTS KWI": "KWI",
    "STTS KBP": "KBP",
    "STTS ABJ": "ABJ",
    "STTS CKY": "CKY",
    "STTS OUA": "OUA",
    "STTS MCT": "MCT",
    "STTS TLV": "TLV",
    "STTS BEY": "BEY",
    "STTS UZK": "UZK",
    "STTS ACC": "ACC",
    "STTS BKO": "BKO",
    "STTS NDJ": "NDJ",
    "STTS NKC": "NKC",
    "STTS NIM": "NIM",
    "STTS FNA": "FNA",
    "STTS FRU": "FRU",
    "STTS LFW": "LFW",
    "STTS BJL": "BJL",
    "STTS CMB": "CMB",
    "STTS KTM": "KTM",
    "STTS IND": "IND",
    "STTS ALG": "ALG",
}

monthly_df["Branch Name"] = (
    monthly_df["Branch Name"]
    .replace(branch_short)
)

#Create Pivot

monthly_report = pd.pivot_table(
    monthly_df,
    index="Date Convert",
    columns="Branch Name",
    values="Sale (€)",
    aggfunc="sum",
    fill_value=0
)
monthly_report.index = pd.to_datetime(monthly_report.index)
monthly_report.index = (
    pd.to_datetime(monthly_report.index)
    .strftime("%d-%b")
    .str.lstrip("0")
)
monthly_report.index.name = "DATE"

# Docs Column

docs = (
    monthly_df
    .assign(
        DATE=monthly_df["Date Convert"]
            .dt.strftime("%d-%b")
            .str.lstrip("0")
    )
    .groupby("DATE")
    .size()
)

monthly_report["Docs"] = docs.reindex(
    monthly_report.index,
    fill_value=0
)
monthly_report["TOTAL"] = monthly_report.sum(axis=1)

# TOTAL

# Remove all zero columns

monthly_report = monthly_report.loc[
    :,
    (monthly_report != 0).any(axis=0)
]

# Recreate branch list AFTER removing zero columns

branch_cols = [
    c for c in monthly_report.columns
    if c != "Docs"
]

# Sort branches by total sales

branch_totals = monthly_report[branch_cols].sum()

sorted_branches = (
    branch_totals
    .sort_values(ascending=False)
    .index
    .tolist()
)


monthly_report = monthly_report[
    sorted_branches
    + ["Docs", "TOTAL"]
]


branch_cols = [
    c
    for c in monthly_report.columns
    if c not in ["Docs", "TOTAL"]
]

# TOTAL Row
monthly_report.loc["TOTAL"] = monthly_report.sum()

# Fix Docs total
monthly_report.loc["TOTAL", "Docs"] = monthly_report.loc[
    monthly_report.index != "TOTAL",
    "Docs"
].sum()

# Fix TOTAL column
monthly_report.loc["TOTAL", "TOTAL"] = monthly_report.loc[
    monthly_report.index != "TOTAL",
    branch_cols
].sum(axis=1).sum() + monthly_report.loc[
    monthly_report.index != "TOTAL",
    "Docs"
].sum()

display_df = monthly_report.copy()

display_df = display_df.round(0)

display_df = display_df.fillna(0)

#display_df = display_df.astype(int)

# -------------------------
# Theme Colors
# -------------------------

if report_type == "ALL":
    header_color = "#215C98"
    data_color   = "#DAE9F8"
    date_color   = "#DAE9F8"

elif report_type == "HQ":
    header_color = "#3C7D22"
    data_color   = "#DAF2D0"
    date_color   = "#B5E6A2"

else:   # B2B
    header_color = "#BE5014"
    data_color   = "#F1A983"
    date_color   = "#F1A983"

#st.write(report_type)
#st.write(header_color)
#st.write(data_color)
#st.write(date_color)

st.write(display_df.columns.tolist())
st.write(display_df.head())

styled_report = (
    display_df.style

    # Number Format
    .format("{:,.0f}")

    # Entire data area
    .set_properties(**{
        "background-color": data_color,
        "border": "0.8px solid #404040",
        "color": "black",
        "text-align": "center"
    })

    # Docs & TOTAL column bold
    .set_properties(
        subset=["Docs"],
        **{
            "font-weight": "bold"
        }
    )

    # Header + Index + TOTAL
    .set_table_styles([

        # Column Headers
        {
            "selector": "th.col_heading",
            "props": [
                ("background-color", header_color),
                ("color", "white"),
                ("font-weight", "bold"),
                ("border", "1px solid black")
            ]
        },

        # DATE column (index)
        {
            "selector": "th.row_heading",
            "props": [
                ("background-color", date_color),
                ("color", "black"),
                ("font-weight", "normal"),
                ("border", "1px solid black")
            ]
        },

        # Top Left Cell
        {
            "selector": "th.blank",
            "props": [
                ("background-color", header_color),
                ("color", "white"),
                ("font-weight", "bold"),
                ("border", "1px solid black")
            ]
        }

    ])

    # TOTAL Row
    .apply(
        lambda row: [
            (
                f"background-color:{header_color};"
                "color:white;"
                "font-weight:bold;"
                "border:1px solid black;"
            )
            if row.name == "TOTAL"
            else ""
            for _ in row
        ],
        axis=1
    )
    .apply_index(
        lambda x: [
            (
                f"background-color:{header_color};"
                "color:white;"
                "font-weight:bold;"
                "border:1px solid black;"
            )
            if v == "TOTAL"
            else (
                f"background-color:{date_color};"
                "color:black;"
                "font-weight:normal;"
                "border:1px solid black;"
            )
            for v in x
        ],
        axis=0
    )
    .set_table_styles([
        {
            "selector": "th.row_heading.level0",
            "props": [
                ("background-color", date_color),
                ("color", "black"),
                ("font-weight", "normal"),
                ("border", "1px solid black")
            ]
        }
    ], overwrite=False)

    # Docs & TOTAL column bold
    .map(
        lambda x: "font-weight:bold;",
        subset=["Docs"]
    )
)

#from streamlit.components.v1 import html
#
#display_df = monthly_report.reset_index()
#
#display_df.rename(
#    columns={"index": "DATE"},
#    inplace=True
#)
#
#html(
#    render_html_table(
#        display_df,
#        report_type
#    ),
#    height=900,
#    scrolling=True
#)


#st.dataframe(
#    styled_report,
##    use_container_width=True,
#    height=650
#)


display_df = monthly_report.copy()
display_df.index.name = ""

display_df = display_df.reset_index()
display_df.rename(columns={"index": "DATE"}, inplace=True)
st.table(styled_report)


# ==========================
# BRANCH PERFORMANCE MATRIX
# ==========================

st.markdown("## 🔥 Branch Performance Matrix")

matrix_df = (
    filtered_df.groupby(
        ["Branch Name", "Month Number"]
    )["Sale (€)"]
    .sum()
    .reset_index()
)

sales_pivot = matrix_df.pivot(
    index="Branch Name",
    columns="Month Number",
    values="Sale (€)"
).fillna(0)

sales_pivot = sales_pivot.reindex(
    columns=range(1,13),
    fill_value=0
)

report_df = pd.DataFrame(
    index=sales_pivot.index
)

month_names = {
    1:"Jan",
    2:"Feb",
    3:"Mar",
    4:"Apr",
    5:"May",
    6:"Jun",
    7:"Jul",
    8:"Aug",
    9:"Sep",
    10:"Oct",
    11:"Nov",
    12:"Dec"
}

max_month = int(
    filtered_df["Month Number"].max()
)

report_df = pd.DataFrame(
    index=sales_pivot.index
)

# Jan Sales first
report_df["Jan Sales"] = sales_pivot[1]

for m in range(2, max_month + 1):

    report_df[
        f"{month_names[m]} Sales"
    ] = sales_pivot[m]

    previous_month = sales_pivot[m-1]

    growth = (
        (sales_pivot[m] - previous_month)
        /
        previous_month.replace(
            0,
            float("nan")
        )
    ) * 100

    report_df[
        f"{month_names[m]} Growth %"
    ] = pd.to_numeric(
        growth,
        errors="coerce"
    ).round(1)

report_df["Total (Euro)"] = (
    sales_pivot.sum(axis=1)
)

# TOTAL Row

total_row = {}

for col in report_df.columns:

    if "Growth" in col:

        month_name = col.replace(
            " Growth %",
            ""
        )

        month_num = None

        for k, v in month_names.items():

            if v == month_name:
                month_num = k
                break

        if month_num and month_num > 1:

            prev_total = sales_pivot[
                month_num - 1
            ].sum()

            curr_total = sales_pivot[
                month_num
            ].sum()

            if prev_total != 0:

                total_row[col] = round(
                    (
                        (curr_total - prev_total)
                        / prev_total
                    ) * 100,
                    1
                )

            else:
                total_row[col] = ""

        else:
            total_row[col] = ""

    else:

        total_row[col] = report_df[
            col
        ].sum()



# Sort branches first

report_df = report_df.sort_values(
    "Total (Euro)",
    ascending=False
)

# Add TOTAL row after sorting

report_df = report_df.sort_values(
    "Total (Euro)",
    ascending=False
)

report_df.loc["TOTAL"] = total_row

# ==========================
# MATRIX VIEW FILTER
# ==========================

view_mode = st.radio(
    "View",
    [
        "Top 10",
        "Top 20",
        "All Branches"
    ],
    horizontal=True,
    index=1,
    key="matrix_view"
)

if view_mode == "Top 10":

    top_rows = report_df.drop(
        "TOTAL",
        errors="ignore"
    ).head(10)

    report_df_display = pd.concat(
        [
            top_rows,
            report_df.loc[["TOTAL"]]
        ]
    )

elif view_mode == "Top 20":

    top_rows = report_df.drop(
        "TOTAL",
        errors="ignore"
    ).head(20)

    report_df_display = pd.concat(
        [
            top_rows,
            report_df.loc[["TOTAL"]]
        ]
    )

else:

    report_df_display = report_df

display_df = report_df_display.copy()

# ==========================
# COLUMN LISTS
# ==========================

sales_cols = [
    col
    for col in report_df.columns
    if "Sales" in col
    or col == "Total (Euro)"
]

growth_cols = [
    col
    for col in report_df.columns
    if "Growth" in col
]

# Format sales columns

for col in sales_cols:

    if col == "Total (Euro)":

        display_df[col] = (
            pd.to_numeric(
                display_df[col],
                errors="coerce"
            )
            .fillna(0)
            .apply(
                lambda x: f"{x:,.0f}"
            )
        )

    else:

        display_df[col] = (
            pd.to_numeric(
                display_df[col],
                errors="coerce"
            )
            .fillna(0)
            .apply(
                lambda x: f"{x:,.0f}"
            )
        )

# Format growth columns

for col in growth_cols:

    display_df[col] = (
        pd.to_numeric(
            display_df[col],
            errors="coerce"
        )
        .apply(
            lambda x:
            "-"
            if pd.isna(x)
            else f"{x:.1f}%"
        )
    )

def growth_color(val):

    try:

        if val == "-":
            return ""

        num = float(
            str(val)
            .replace("%","")
        )

        if num > 0:
            return (
                "background-color:"
                "#c6efce;"
                "color:#006100;"
            )

        if num < 0:
            return (
                "background-color:"
                "#ffc7ce;"
                "color:#9c0006;"
            )

    except:
        pass

    return ""

def total_row_style(row):

    if row.name == "TOTAL":

        return [
            "background-color:#0b2e59;color:white;font-weight:bold"
        ] * len(row)

    return [""] * len(row)


styled_df = (
    display_df.style
    .map(
        growth_color,
        subset=growth_cols
    )
    .apply(
        total_row_style,
        axis=1
    )
)

if report_type == "HQ":
    st.markdown(
        "<h3 style='color:#548235;'>🟩 HQ SALES REPORT</h3>",
        unsafe_allow_html=True
    )

elif report_type == "B2B":
    st.markdown(
        "<h3 style='color:#C55A11;'>🟧 B2B SALES REPORT</h3>",
        unsafe_allow_html=True
    )

# ALL report -> No title

st.dataframe(
    styled_df,
    use_container_width=False,
    height=700
)

# ==========================
# EXCEL EXPORT
# ==========================

def create_matrix_excel(export_df):

    wb = Workbook()
    ws = wb.active
    ws.title = "Branch Matrix"

    # Headers
    for col_num, col_name in enumerate(export_df.columns, 1):

        cell = ws.cell(
            row=1,
            column=col_num + 1
        )

        cell.value = col_name

        # Header Color
        if report_type == "ALL":
            header_color = "1F4E78"      # Blue

        elif report_type == "HQ":  
            header_color = "548235"      # Green

        else:
            header_color = "C55A11"      # Orange  

        cell.fill = PatternFill(   
            "solid",
            fgColor=header_color
        )



        cell.font = Font(
            color="FFFFFF",
            bold=True
        )

    # Branch names
    for row_num, idx in enumerate(
        export_df.index,
        2
    ):

        ws.cell(
            row=row_num,
            column=1
        ).value = idx

    # Data
    for row_num, (_, row) in enumerate(
        export_df.iterrows(),
        2
    ):

        for col_num, value in enumerate(
            row,
            2
        ):

            cell = ws.cell(
                row=row_num,
                column=col_num
            )

            # Data Row Color
            if report_type == "ALL":
                row_color = "DCE6F1"      # Light Blue

            elif report_type == "HQ":
                row_color = "E2F0D9"      # Light Green

            else:
                row_color = "FCE4D6"      # Light Orange

            cell.fill = PatternFill(
                "solid",
                fgColor=row_color
            )       

            
            if value == 0:
                    cell.value = "-"

            elif isinstance(value, (int, float)): 
                cell.value = round(value)
                cell.number_format = '#,##0'
            else:
                cell.value = value
    


            # Growth colors
            if "Growth" in export_df.columns[
                col_num - 2
            ]:

                try:

                    num = float(
                        str(value)
                        .replace("%", "")
                    )

                    if num > 0:

                        cell.fill = PatternFill(
                            "solid",
                            fgColor="C6EFCE"
                        )

                    elif num < 0:

                        cell.fill = PatternFill(
                            "solid",
                            fgColor="FFC7CE"
                        )

                except:
                    pass

    # TOTAL row
    for row in ws.iter_rows():

        if row[0].value == "TOTAL":

            for cell in row:

                cell.fill = PatternFill(
                    "solid",
                    fgColor=header_color
                )

                cell.font = Font(
                    color="FFFFFF",
                    bold=True
                )

    ws.freeze_panes = "B2"

    # Border for all cells
    thin = Side(
        style="thin",
        color="D9D9D9"
    )

    border = Border(
        left=thin,
        right=thin,
        top=thin,
        bottom=thin
    )
    for row in ws.iter_rows():
        for cell in row:
            cell.border = border


    # Make TOTAL column bold
    grand_total_col = ws.max_column
    for row in range(2, ws.max_row + 1):
        ws.cell(
            row=row,
            column=grand_total_col).font = Font(
                bold=True
            )
        
    # Make DATE column bold

    for row in range(2, ws.max_row + 1):
        ws.cell(
        row=row,
        column=1
    ).font = Font(
        bold=True
    )

    for column in ws.columns:
        length = 0
        letter = get_column_letter(column[0].column)

        for cell in column:
            try:
                length = max(length, len(str(cell.value)))
            except:
                pass

        ws.column_dimensions[letter].width = length + 3

    excel_file = BytesIO()

    wb.save(excel_file)

    excel_file.seek(0)

    return excel_file

excel_data = create_matrix_excel(
    display_df
)

st.download_button(
    "📥 Download Matrix Report",
    data=excel_data,
    file_name="Branch_Performance_Matrix.xlsx",
    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
)

# ==========================
# BRANCH WISE DETAIL
# ==========================

if not pdf_mode:

    st.markdown("## 📋 Branch Wise Detail")

    branch_table = (
        filtered_df.groupby("Branch Name")["Sale (€)"]
        .sum()
        .reset_index()
        .sort_values(
            by="Sale (€)",
            ascending=False
        )
    )

    branch_table["Sale (€)"] = (
        branch_table["Sale (€)"]
        .apply(lambda x: f"{x:,.0f}")
    )

    csv = branch_table.to_csv(index=False)

    col1, col2 = st.columns([4, 1])

    with col1:
        search_branch = st.text_input(
            "🔍 Search Branch",
            placeholder="STTS MCT"
        )

    with col2:
        st.download_button(
            "📥 Download",
            csv,
            "branch_report.csv",
            "text/csv"
        )

    if search_branch:
        search_result = branch_table[
            branch_table["Branch Name"]
            .str.contains(search_branch, case=False, na=False)
        ]

        if not search_result.empty:
            st.success(f"Found {len(search_result)} branch(es)")
            st.dataframe(
                search_result,
                use_container_width=True,
                hide_index=True
            )

    st.dataframe(
        branch_table,
        use_container_width=True,
        hide_index=True
    )
